"""
Patch server.py to add Excel export routes and room type inventory.
Run once: python add_excel_routes.py
"""
content = open(r'C:\Users\bryan\WorkBuddy\20260426211044\hotel-app\server.py', 'r', encoding='utf-8').read()

excel_code = '''
# ─── Excel Export ─────────────────────────────────────────────────────────────
# 酒店房型配置（固定库存）
ROOM_TYPE_INVENTORY = {
    "特惠大床房":   {"total": 2},
    "精选大床房":   {"total": 12},
    "豪华大床房":   {"total": 14},
    "城堡亲子房":  {"total": 4},
    "麻将双床房":   {"total": 7},
    "麻将套房":     {"total": 2},
    "圆床房":       {"total": 1},
}

@app.route("/api/reports/export/daily", methods=["GET"])
@boss_required
def api_export_daily():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 未安装"}), 500

    year = request.args.get("year")
    month = request.args.get("month")
    db = get_db()
    if year and month:
        rows = db.execute(
            "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date ASC",
            (f"{year}-{month.zfill(2)}%",)
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM daily_reports ORDER BY date ASC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "日报明细"

    HDR_FILL = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUB_FILL = PatternFill("solid", fgColor="E8EDF5")
    SUB_FONT = Font(bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=14, color="1A1A2E")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    money_cols = {4, 6, 8, 10, 12, 14, 16, 18, 19, 24}

    ws.merge_cells("A1:X1")
    title_text = "华悦艺术酒店（冠华店）日报明细"
    if year and month:
        title_text += f" {year}年{int(month):02d}月"
    ws["A1"] = title_text
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws.row_dimensions[1].height = 28

    headers = [
        "日期", "班次",
        "美团间", "美团收入", "携程间", "携程收入",
        "飞猪间", "飞猪收入", "抖音间", "抖音收入",
        "微信间", "微信收入", "现金间", "现金收入",
        "支付宝间", "支付宝收入",
        "停车票", "停车收入", "税",
        "开房间", "入住率%", "RevPGR", "营业额",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = center; cell.border = border
    ws.row_dimensions[3].height = 22

    for ri, r in enumerate(rows, 4):
        vals = [
            r["date"], r["shift"] or "",
            r["meituan_rooms"], r["meituan_income"],
            r["ctrip_rooms"], r["ctrip_income"],
            r["fliggy_rooms"], r["fliggy_income"],
            r["douyin_rooms"], r["douyin_income"],
            r["wechat_rooms"], r["wechat_income"],
            r["cash_rooms"], r["cash_income"],
            r["alipay_rooms"], r["alipay_income"],
            r["parking_tickets"], r["parking_income"], r["tax"],
            r["total_rooms"], r["occupancy_rate"], r["revpgr"], r["total_income"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = border
            if col == 1:
                cell.value = r["date"]
                cell.font = Font(bold=True, color="0F3460")
                cell.alignment = center
            elif col == 2:
                cell.value = r["shift"] or ""
                cell.alignment = center
            elif col in money_cols:
                cell.value = round(float(v or 0), 2)
                cell.number_format = '"¥"#,##0.00'
                cell.font = Font(color="E94560", bold=True)
                cell.alignment = center
            elif col == 21:
                occ = float(v or 0)
                cell.value = round(occ, 1)
                cell.number_format = "0.0"
                cell.alignment = center
                cell.font = Font(color="2E7D32" if occ >= 70 else "E65100", bold=True)
            else:
                cell.value = v if v is not None else 0
                cell.alignment = center

    last_row = 4 + len(rows)
    ws.cell(row=last_row, column=1, value="汇总").fill = SUB_FILL
    ws.cell(row=last_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=last_row, column=1).alignment = center; ws.cell(row=last_row, column=1).border = border
    sum_cols = {3: "meituan_rooms", 4: "meituan_income", 5: "ctrip_rooms", 6: "ctrip_income",
                 7: "fliggy_rooms", 8: "fliggy_income", 9: "douyin_rooms", 10: "douyin_income",
                 11: "wechat_rooms", 12: "wechat_income", 13: "cash_rooms", 14: "cash_income",
                 15: "alipay_rooms", 16: "alipay_income", 17: "parking_tickets", 18: "parking_income",
                 19: "tax", 20: "total_rooms", 23: "total_income"}
    for col, key in sum_cols.items():
        val = sum(float(r[key] or 0) for r in rows)
        cell = ws.cell(row=last_row, column=col, value=round(val, 2))
        cell.fill = SUB_FILL; cell.font = SUB_FONT
        cell.alignment = center; cell.border = border
        if col in money_cols:
            cell.number_format = '"¥"#,##0.00'

    for i, w in enumerate([13, 7, 7, 12, 7, 12, 7, 12, 7, 12, 7, 12, 7, 12, 8, 12, 8, 12, 8, 8, 8, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"日报明细_{year or '全部'}_{month or ''}".strip("_") + ".xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/api/reports/export/roomtypes", methods=["GET"])
@boss_required
def api_export_roomtypes():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 未安装"}), 500

    year = request.args.get("year")
    month = request.args.get("month")
    db = get_db()
    if year and month:
        rows = db.execute(
            "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date ASC",
            (f"{year}-{month.zfill(2)}%",)
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM daily_reports ORDER BY date ASC").fetchall()

    reports = []
    for r in rows:
        d = dict(r)
        if d.get("room_types"):
            try:
                d["room_types"] = json.loads(d["room_types"])
            except:
                d["room_types"] = {}
        else:
            d["room_types"] = {}
        reports.append(d)

    wb = Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
    SUB_FILL  = PatternFill("solid", fgColor="E8EDF5")
    SUB_FONT  = Font(bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=14, color="1A1A2E")
    GOOD_FILL = PatternFill("solid", fgColor="E8F5E9")
    WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
    BAD_FILL  = PatternFill("solid", fgColor="FFEBEE")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    period = f"{year}年{int(month):02d}月" if (year and month) else "全部时间"
    days = max(len(reports), 1)

    # Sheet 1: 房型汇总
    ws1 = wb.active
    ws1.title = "房型汇总"
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"华悦艺术酒店（冠华店）房型经营汇总  {period}"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = center
    ws1["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws1.row_dimensions[1].height = 28

    hdrs = ["房型", "总房间数", "累计开房数", "开房率", "日均开房", "总营业额", "平均房价", "开房率", "经营评价"]
    for col, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = center; cell.border = border
    ws1.row_dimensions[3].height = 22

    totals = {name: {"rooms": 0, "income": 0} for name in ROOM_TYPE_INVENTORY}
    for r in reports:
        rt = r.get("room_types") or {}
        for name in ROOM_TYPE_INVENTORY:
            if name in rt:
                totals[name]["rooms"]  += int(rt[name].get("rooms", 0))
                totals[name]["income"] += float(rt[name].get("income", 0))

    total_inv = sum(v["total"] for v in ROOM_TYPE_INVENTORY.values())
    total_occ = sum(v["rooms"] for v in totals.values())
    total_inc = sum(v["income"] for v in totals.values())

    for ri, (name, inv) in enumerate(ROOM_TYPE_INVENTORY.items(), 4):
        t = totals[name]
        occ = t["rooms"]; inc = t["income"]
        occ_rate = (occ / (inv["total"] * days) * 100) if days > 0 else 0
        avg_p = inc / occ if occ > 0 else 0
        if occ == 0:
            rating, rfill = "无数据", BAD_FILL
        elif occ_rate >= 80:
            rating, rfill = "优秀", GOOD_FILL
        elif occ_rate >= 50:
            rating, rfill = "良好", PatternFill("solid", fgColor="E3F2FD")
        else:
            rating, rfill = "待提升", WARN_FILL

        vals = [name, inv["total"], occ, f"{occ_rate:.1f}%",
                f"{occ/days:.1f}", f"¥{inc:,.2f}", f"¥{avg_p:.2f}", f"{occ_rate:.1f}%", rating]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=col, value=v)
            cell.border = border; cell.alignment = center
            if col == 9:
                cell.fill = rfill; cell.font = Font(bold=True)
            elif col == 1:
                cell.font = Font(bold=True, color="0F3460")
            elif col in (2, 3):
                cell.font = Font(bold=True)
        ws1.row_dimensions[ri].height = 20

    last = 4 + len(ROOM_TYPE_INVENTORY)
    ws1.cell(row=last, column=1, value="合计").fill = SUB_FILL
    ws1.cell(row=last, column=1).font = Font(bold=True, size=11)
    ws1.cell(row=last, column=1).alignment = center; ws1.cell(row=last, column=1).border = border
    ws1.cell(row=last, column=2, value=total_inv).fill = SUB_FILL
    ws1.cell(row=last, column=2).font = Font(bold=True); ws1.cell(row=last, column=2).alignment = center; ws1.cell(row=last, column=2).border = border
    ws1.cell(row=last, column=3, value=total_occ).fill = SUB_FILL
    ws1.cell(row=last, column=3).font = Font(bold=True); ws1.cell(row=last, column=3).alignment = center; ws1.cell(row=last, column=3).border = border
    ws1.cell(row=last, column=6, value=f"¥{total_inc:,.2f}").fill = SUB_FILL
    ws1.cell(row=last, column=6).font = Font(bold=True, color="E94560"); ws1.cell(row=last, column=6).alignment = center; ws1.cell(row=last, column=6).border = border
    avg_all = total_inc / total_occ if total_occ > 0 else 0
    ws1.cell(row=last, column=7, value=f"¥{avg_all:.2f}").fill = SUB_FILL
    ws1.cell(row=last, column=7).font = Font(bold=True); ws1.cell(row=last, column=7).alignment = center; ws1.cell(row=last, column=7).border = border
    for i, w in enumerate([14, 10, 12, 10, 10, 14, 12, 10, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A4"

    # Sheet 2: 每日明细
    ws2 = wb.create_sheet("每日房型明细")
    ws2.merge_cells("A1:K1")
    ws2["A1"] = f"华悦艺术酒店（冠华店）每日房型明细  {period}"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = center
    ws2["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws2.row_dimensions[1].height = 28

    daily_hdrs = ["日期", "班次"] + list(ROOM_TYPE_INVENTORY.keys()) + ["当日总开房", "当日营业额"]
    for col, h in enumerate(daily_hdrs, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = center; cell.border = border
    ws2.row_dimensions[3].height = 22

    for ri, r in enumerate(reports, 4):
        rt = r.get("room_types") or {}
        row_vals = [r["date"], r["shift"] or ""]
        day_total = 0
        for name in ROOM_TYPE_INVENTORY:
            rooms = int(rt.get(name, {}).get("rooms", 0))
            row_vals.append(rooms); day_total += rooms
        row_vals.append(day_total)
        row_vals.append(round(float(r.get("total_income", 0) or 0), 2))
        for col, v in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=col, value=v)
            cell.border = border; cell.alignment = center
            if col == 1:
                cell.font = Font(bold=True, color="0F3460")
            elif col == 11:
                cell.number_format = '"¥"#,##0.00'
                cell.font = Font(color="E94560", bold=True)
            elif 3 <= col <= 9:
                occ = int(v or 0)
                cell.fill = PatternFill("solid", fgColor="E8F5E9") if occ > 0 else PatternFill("solid", fgColor="FAFAFA")
                if occ == 0:
                    cell.font = Font(color="CCCCCC")

    for i, w in enumerate([13, 7, 10, 10, 10, 10, 10, 10, 10, 12, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "C4"

    # Sheet 3: 经营分析
    ws3 = wb.create_sheet("经营分析")
    ws3.merge_cells("A1:B1")
    ws3["A1"] = f"房型经营分析报告  {period}"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = center
    ws3["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws3.row_dimensions[1].height = 28
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 65

    analyses = []
    for name, inv in ROOM_TYPE_INVENTORY.items():
        t = totals[name]
        occ = t["rooms"]; inc = t["income"]
        occ_rate = (occ / (inv["total"] * days) * 100) if days > 0 else 0
        avg_p = inc / occ if occ > 0 else 0
        analyses.append({"name": name, "inv": inv["total"], "occ": occ,
                          "inc": inc, "rate": occ_rate, "avg": avg_p})

    best_inc  = max(analyses, key=lambda x: x["inc"]) if analyses else None
    best_rate = max(analyses, key=lambda x: x["rate"]) if analyses else None
    worst = min([x for x in analyses if x["occ"] > 0], key=lambda x: x["rate"], default=None)

    rows3 = [
        ("数据概览", f"统计周期：{period}  |  共 {len(reports)} 天  |  客房总数：{total_inv} 间"),
        ("汇总", f"累计开房：{total_occ} 间夜  |  房型总收入：¥{total_inc:,.2f}  |  综合均价：{'¥'+str(round(total_inc/total_occ,2)) if total_occ > 0 else '¥0'}"),
        ("收入冠军", f'{best_inc["name"]}：总收入 ¥{best_inc["inc"]:,.2f}，累计开房 {best_inc["occ"]} 间夜，均价 ¥{best_inc["avg"]:.2f}'),
        ("开房率冠军", f'{best_rate["name"]}：开房率 {best_rate["rate"]:.1f}%，累计开房 {best_rate["occ"]} 间夜'),
        ("待提升房型", f'{worst["name"] if worst else "无"}：开房率 {worst["rate"]:.1f+"%" if worst else "0%"}，建议关注价格策略或房型配置'),
        ("经营建议", "高收入房型保障流量支持；低入住率房型建议适度调价或优化套餐组合，提升整体收益"),
    ]
    for ri, (label, content_txt) in enumerate(rows3, 3):
        c1 = ws3.cell(row=ri, column=1, value=label)
        c1.font = Font(bold=True, size=11, color="0F3460")
        c1.fill = PatternFill("solid", fgColor="E8EDF5")
        c1.alignment = center; c1.border = border
        c2 = ws3.cell(row=ri, column=2, value=content_txt)
        c2.font = Font(size=11)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        c2.border = border
        ws3.row_dimensions[ri].height = 32

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"房型汇总_{year or '全部'}_{month or ''}".strip("_") + ".xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


'''

# Insert before "# ─── Frontend"
frontend_marker = 'Frontend ───────────────────────────────────────────────────────────────'
idx = content.index(frontend_marker)
new_content = content[:idx] + excel_code + content[idx:]
open(r'C:\Users\bryan\WorkBuddy\20260426211044\hotel-app\server.py', 'w', encoding='utf-8').write(new_content)
print("Done! Inserted Excel routes at index", idx)
