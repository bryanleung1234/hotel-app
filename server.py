# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

"""
酒店日报管理系统 - Python 后端
支持 SQLite（本地）和 PostgreSQL（Render云端，通过 DATABASE_URL 环境变量切换）
依赖: pip install flask pyjwt
PostgreSQL: pip install psycopg2-binary
"""
import os
import jwt
import datetime
import json
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, g, send_file

app = Flask(__name__, static_folder='public', static_url_path='')
app.config['SECRET_KEY'] = 'hotel-report-secret-key-2024'
JWT_SECRET = app.config['SECRET_KEY']

# 自动检测数据库模式：有 DATABASE_URL 用 PostgreSQL，否则用本地 SQLite
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_POSTGRES = bool(DATABASE_URL)

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    print(f'[DB] 使用 PostgreSQL')
else:
    import sqlite3
    app.config['DATABASE'] = os.path.join(os.path.dirname(__file__), 'hotel.db')
    print(f'[DB] 使用 SQLite: {app.config["DATABASE"]}')

# --- Database ----------------------------------------------------------------
def get_db():
    if 'db' not in g:
        if USE_POSTGRES:
            conn = psycopg2.connect(DATABASE_URL)
            conn.autocommit = False
            g.db = conn
        else:
            import sqlite3
            g.db = sqlite3.connect(app.config['DATABASE'])
            g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        try:
            db.close()
        except Exception:
            pass

def db_execute(db, sql, params=()):
    """统一执行 SQL，自动处理 SQLite/PostgreSQL 占位符差异"""
    if USE_POSTGRES:
        # PostgreSQL 用 %s 占位符
        sql_pg = sql.replace('?', '%s')
        # 处理 INSERT OR IGNORE → INSERT ... ON CONFLICT DO NOTHING
        sql_pg = sql_pg.replace('INSERT OR IGNORE INTO', 'INSERT INTO').replace(
            'INSERT OR IGNORE', 'INSERT')
        if 'INSERT INTO users' in sql_pg and 'OR IGNORE' not in sql and 'IGNORE' not in sql_pg:
            pass
        if 'INSERT OR REPLACE INTO monthly_cache' in sql:
            sql_pg = sql_pg.replace('INSERT OR REPLACE INTO monthly_cache',
                'INSERT INTO monthly_cache')
            sql_pg = sql_pg.rstrip(')') + ') ON CONFLICT (year, month) DO UPDATE SET data=EXCLUDED.data, updated_at=EXCLUDED.updated_at'
        if 'INSERT OR IGNORE INTO users' in sql:
            sql_pg = sql.replace('?', '%s').replace('INSERT OR IGNORE INTO users',
                'INSERT INTO users') + ' ON CONFLICT (phone) DO NOTHING'
        if 'INSERT OR REPLACE INTO daily_reports' in sql:
            sql_pg = sql.replace('?', '%s').replace('INSERT OR REPLACE INTO daily_reports',
                'INSERT INTO daily_reports')
            sql_pg = sql_pg.rstrip(')') + ''') ON CONFLICT (date) DO UPDATE SET
                shift=EXCLUDED.shift, meituan_rooms=EXCLUDED.meituan_rooms,
                ctrip_rooms=EXCLUDED.ctrip_rooms, fliggy_rooms=EXCLUDED.fliggy_rooms,
                douyin_rooms=EXCLUDED.douyin_rooms, wechat_rooms=EXCLUDED.wechat_rooms,
                cash_rooms=EXCLUDED.cash_rooms, alipay_rooms=EXCLUDED.alipay_rooms,
                meituan_income=EXCLUDED.meituan_income, ctrip_income=EXCLUDED.ctrip_income,
                fliggy_income=EXCLUDED.fliggy_income, douyin_income=EXCLUDED.douyin_income,
                wechat_income=EXCLUDED.wechat_income, cash_income=EXCLUDED.cash_income,
                alipay_income=EXCLUDED.alipay_income, parking_tickets=EXCLUDED.parking_tickets,
                parking_income=EXCLUDED.parking_income, tax=EXCLUDED.tax,
                total_rooms=EXCLUDED.total_rooms, avg_price=EXCLUDED.avg_price,
                occupancy_rate=EXCLUDED.occupancy_rate, revpgr=EXCLUDED.revpgr,
                total_income=EXCLUDED.total_income, note=EXCLUDED.note,
                room_types=EXCLUDED.room_types, uploaded_by=EXCLUDED.uploaded_by'''
        cur = db.cursor()
        cur.execute(sql_pg, params)
        return PGCursorWrapper(cur)
    else:
        return db.execute(sql, params)

class PGCursorWrapper:
    """让 psycopg2 cursor 像 sqlite3 Row 一样可以用字段名访问"""
    def __init__(self, cur):
        self._cur = cur
    def fetchall(self):
        rows = self._cur.fetchall()
        if not rows:
            return []
        desc = self._cur.description
        if not desc:
            return []
        cols = [d[0] for d in desc]
        return [dict(zip(cols, row)) for row in rows]
    def fetchone(self):
        row = self._cur.fetchone()
        if not row:
            return None
        desc = self._cur.description
        if not desc:
            return None
        cols = [d[0] for d in desc]
        return dict(zip(cols, row))
    @property
    def rowcount(self):
        return self._cur.rowcount

def init_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute('''
          CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            phone TEXT UNIQUE NOT NULL,
            name TEXT,
            role TEXT NOT NULL DEFAULT 'staff',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
          )
        ''')
        cur.execute('''
          CREATE TABLE IF NOT EXISTS daily_reports (
            id SERIAL PRIMARY KEY,
            date TEXT NOT NULL UNIQUE,
            shift TEXT,
            meituan_rooms INTEGER DEFAULT 0,
            ctrip_rooms INTEGER DEFAULT 0,
            fliggy_rooms INTEGER DEFAULT 0,
            douyin_rooms INTEGER DEFAULT 0,
            wechat_rooms INTEGER DEFAULT 0,
            cash_rooms INTEGER DEFAULT 0,
            alipay_rooms INTEGER DEFAULT 0,
            meituan_income REAL DEFAULT 0,
            ctrip_income REAL DEFAULT 0,
            fliggy_income REAL DEFAULT 0,
            douyin_income REAL DEFAULT 0,
            wechat_income REAL DEFAULT 0,
            cash_income REAL DEFAULT 0,
            alipay_income REAL DEFAULT 0,
            parking_tickets INTEGER DEFAULT 0,
            parking_income REAL DEFAULT 0,
            tax REAL DEFAULT 0,
            total_rooms INTEGER DEFAULT 0,
            avg_price REAL DEFAULT 0,
            occupancy_rate REAL DEFAULT 0,
            revpgr REAL DEFAULT 0,
            total_income REAL DEFAULT 0,
            note TEXT,
            room_types TEXT,
            uploaded_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
          )
        ''')
        cur.execute('''
          CREATE TABLE IF NOT EXISTS monthly_cache (
            id SERIAL PRIMARY KEY,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            data TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(year, month)
          )
        ''')
        cur.execute('''
          CREATE TABLE IF NOT EXISTS login_log (
            id SERIAL PRIMARY KEY,
            phone TEXT, role TEXT, ip TEXT, ua TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
          )
        ''')
        # Seed users
        for phone, name, role in [
            ('19128957480', '员工', 'staff'),
            ('13802531098', '老板A', 'boss'),
            ('18602032126', '老板B', 'boss'),
        ]:
            cur.execute('INSERT INTO users (phone, name, role) VALUES (%s, %s, %s) ON CONFLICT (phone) DO NOTHING',
                        (phone, name, role))
        conn.commit()
        cur.close()
        conn.close()
    else:
        import sqlite3
        db = sqlite3.connect(app.config['DATABASE'])
        db.executescript('''
          CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT UNIQUE NOT NULL,
            name TEXT,
            role TEXT NOT NULL DEFAULT 'staff',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
          );
          CREATE TABLE IF NOT EXISTS daily_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            shift TEXT,
            meituan_rooms INTEGER DEFAULT 0,
            ctrip_rooms INTEGER DEFAULT 0,
            fliggy_rooms INTEGER DEFAULT 0,
            douyin_rooms INTEGER DEFAULT 0,
            wechat_rooms INTEGER DEFAULT 0,
            cash_rooms INTEGER DEFAULT 0,
            alipay_rooms INTEGER DEFAULT 0,
            meituan_income REAL DEFAULT 0,
            ctrip_income REAL DEFAULT 0,
            fliggy_income REAL DEFAULT 0,
            douyin_income REAL DEFAULT 0,
            wechat_income REAL DEFAULT 0,
            cash_income REAL DEFAULT 0,
            alipay_income REAL DEFAULT 0,
            parking_tickets INTEGER DEFAULT 0,
            parking_income REAL DEFAULT 0,
            tax REAL DEFAULT 0,
            total_rooms INTEGER DEFAULT 0,
            avg_price REAL DEFAULT 0,
            occupancy_rate REAL DEFAULT 0,
            revpgr REAL DEFAULT 0,
            total_income REAL DEFAULT 0,
            note TEXT,
            room_types TEXT,
            uploaded_by TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(date)
          );
          CREATE TABLE IF NOT EXISTS monthly_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            data TEXT NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(year, month)
          );
          CREATE TABLE IF NOT EXISTS login_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT, role TEXT, ip TEXT, ua TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
          );
        ''')
        try:
            db.execute('ALTER TABLE daily_reports ADD COLUMN room_types TEXT')
            db.commit()
        except Exception:
            pass
        for phone, name, role in [
            ('19128957480', '员工', 'staff'),
            ('13802531098', '老板A', 'boss'),
            ('18602032126', '老板B', 'boss'),
        ]:
            db.execute('INSERT OR IGNORE INTO users (phone, name, role) VALUES (?, ?, ?)', (phone, name, role))
        db.commit()
        db.close()
    print('[OK] Database init done')

# --- Auth -------------------------------------------------------------------
def make_token(user):
    payload = {
        'id': user['id'], 'phone': user['phone'],
        'role': user['role'], 'name': user.get('name', ''),
        'exp': datetime.datetime.utcnow() + datetime.timedelta(days=30)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')

def decode_token(token):
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except jwt.ExpiredSignatureError:
        raise ValueError('Token已过期')
    except jwt.InvalidTokenError:
        raise ValueError('无效Token')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'error': '未登录，请先登录'}), 401
        try:
            g.user = decode_token(auth[7:])
        except ValueError as e:
            return jsonify({'error': str(e)}), 401
        return f(*args, **kwargs)
    return decorated

def boss_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if g.user.get('role') != 'boss':
            return jsonify({'error': '权限不足，仅老板可访问'}), 403
        return f(*args, **kwargs)
    return decorated

# --- Routes -----------------------------------------------------------------
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    phone = (data.get('phone') or '').strip()
    code = str(data.get('code') or '').strip()

    if not phone or not code:
        return jsonify({'error': '手机号和验证码必填'}), 400
    if len(code) != 6 or not code.isdigit():
        return jsonify({'error': '验证码为6位数字'}), 400

    db = get_db()
    user = db_execute(db, 'SELECT * FROM users WHERE phone = ?', (phone,)).fetchone()
    if not user:
        return jsonify({'error': '该手机号未注册'}), 401

    db_execute(db,
        'INSERT INTO login_log (phone, role, ip, ua) VALUES (?, ?, ?, ?)',
        (phone, user['role'], request.remote_addr, request.headers.get('User-Agent', ''))
    )
    db.commit()

    user_dict = dict(user)
    token = make_token(user_dict)
    return jsonify({
        'token': token,
        'user': {'id': user_dict['id'], 'phone': user_dict['phone'], 'role': user_dict['role'], 'name': user_dict.get('name', '')}
    })

@app.route('/api/auth/me', methods=['GET'])
@login_required
def api_me():
    return jsonify({'user': g.user})

# --- Upload -----------------------------------------------------------------
@app.route('/api/reports/upload', methods=['POST'])
@login_required
def api_upload():
    data = request.get_json() or {}
    reports = data.get('reports', [])
    if not reports:
        return jsonify({'error': '缺少 reports 数据'}), 400

    db = get_db()
    for r in reports:
        room_types_json = json.dumps(r.get('room_types') or {}, ensure_ascii=False)
        db_execute(db, '''
          INSERT OR REPLACE INTO daily_reports
          (date, shift, meituan_rooms, ctrip_rooms, fliggy_rooms, douyin_rooms,
           wechat_rooms, cash_rooms, alipay_rooms, meituan_income, ctrip_income,
           fliggy_income, douyin_income, wechat_income, cash_income, alipay_income,
           parking_tickets, parking_income, tax, total_rooms, avg_price,
           occupancy_rate, revpgr, total_income, note, room_types, uploaded_by)
          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
          r.get('date'), r.get('shift', ''),
          r.get('meituan_rooms', 0), r.get('ctrip_rooms', 0), r.get('fliggy_rooms', 0),
          r.get('douyin_rooms', 0), r.get('wechat_rooms', 0), r.get('cash_rooms', 0),
          r.get('alipay_rooms', 0), r.get('meituan_income', 0), r.get('ctrip_income', 0),
          r.get('fliggy_income', 0), r.get('douyin_income', 0), r.get('wechat_income', 0),
          r.get('cash_income', 0), r.get('alipay_income', 0), r.get('parking_tickets', 0),
          r.get('parking_income', 0), r.get('tax', 0), r.get('total_rooms', 0),
          r.get('avg_price', 0), r.get('occupancy_rate', 0), r.get('revpgr', 0),
          r.get('total_income', 0), r.get('note', ''), room_types_json, g.user['phone']
        ))
    db.commit()

    # Invalidate monthly cache
    months = list({r['date'][:7] for r in reports if r.get('date')})
    for m in months:
        y, mo = m.split('-')
        db_execute(db, 'DELETE FROM monthly_cache WHERE year = ? AND month = ?', (int(y), int(mo)))
    db.commit()

    return jsonify({'success': True, 'count': len(reports)})

# --- Query ------------------------------------------------------------------
@app.route('/api/reports', methods=['GET'])
@boss_required
def api_reports():
    db = get_db()
    year = request.args.get('year')
    month = request.args.get('month')

    if year and month:
        rows = db_execute(db,
            "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date DESC",
            (f'{year}-{month.zfill(2)}%',)
        ).fetchall()
    elif year:
        rows = db_execute(db,
            "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date DESC",
            (f'{year}%',)
        ).fetchall()
    else:
        rows = db_execute(db, 'SELECT * FROM daily_reports ORDER BY date DESC').fetchall()

    reports_out = []
    for r in rows:
        d = dict(r)
        if d.get('room_types'):
            try:
                d['room_types'] = json.loads(d['room_types'])
            except:
                d['room_types'] = {}
        else:
            d['room_types'] = {}
        reports_out.append(d)
    return jsonify({'reports': reports_out})

@app.route('/api/reports/monthly', methods=['GET'])
@boss_required
def api_monthly():
    year = request.args.get('year')
    month = request.args.get('month')
    if not year or not month:
        return jsonify({'error': '缺少 year 或 month 参数'}), 400

    year, month = int(year), int(month)
    padded = f'{year}-{str(month).zfill(2)}'

    db = get_db()

    # Check cache
    cached = db_execute(db,
        'SELECT data FROM monthly_cache WHERE year = ? AND month = ?', (year, month)
    ).fetchone()
    if cached:
        cached_data = json.loads(cached['data'])
        for r in cached_data.get('reports', []):
            if r.get('room_types') and isinstance(r['room_types'], str):
                try:
                    r['room_types'] = json.loads(r['room_types'])
                except:
                    r['room_types'] = {}
        return jsonify(cached_data)

    # Build from daily reports
    rows = db_execute(db,
        "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date ASC", (padded + '%',)
    ).fetchall()

    if not rows:
        return jsonify({'year': year, 'month': month, 'reports': [], 'summary': None})

    reports = []
    for r in rows:
        d = dict(r)
        # 解析 room_types JSON
        if d.get('room_types'):
            try:
                d['room_types'] = json.loads(d['room_types'])
            except:
                d['room_types'] = {}
        else:
            d['room_types'] = {}
        reports.append(d)
    total_days = len(reports)

    total_rooms = sum(r['total_rooms'] or 0 for r in reports)
    total_income = sum(r['total_income'] or 0 for r in reports)
    avg_income = total_income / total_days
    avg_occ = sum(r['occupancy_rate'] or 0 for r in reports) / total_days
    avg_revpgr = sum(r['revpgr'] or 0 for r in reports) / total_days
    total_tax = sum(r['tax'] or 0 for r in reports)
    total_parking = sum(r['parking_income'] or 0 for r in reports)

    # Channel data
    ch = {k: {'rooms': 0, 'income': 0} for k in ['meituan','ctrip','fliggy','douyin','wechat','cash','alipay']}
    for r in reports:
        for k in ch:
            ch[k]['rooms'] += r.get(f'{k}_rooms', 0) or 0
            ch[k]['income'] += r.get(f'{k}_income', 0) or 0

    # Weekly
    weeks = [[], [], [], [], []]
    for r in reports:
        day = int(r['date'].split('-')[2])
        weeks[min(day // 7, 4)].append(r)
    weekly = [
        {'week': f'第{i+1}周', 'count': len(w),
         'income': sum(r['total_income'] or 0 for r in w),
         'occupancy': sum(r['occupancy_rate'] or 0 for r in w) / len(w) if w else 0}
        for i, w in enumerate(weeks) if w
    ]

    # Daily trend
    trend = [
        {'date': r['date'], 'income': r['total_income'] or 0,
         'occupancy': r['occupancy_rate'] or 0, 'revpgr': r['revpgr'] or 0, 'rooms': r['total_rooms'] or 0}
        for r in reports
    ]

    result = {
        'year': year, 'month': month, 'reportCount': total_days,
        'summary': {
            'totalIncome': total_income, 'avgIncome': avg_income,
            'avgOccupancy': avg_occ, 'avgRevpgr': avg_revpgr,
            'totalTax': total_tax, 'totalParking': total_parking,
            'channelData': ch, 'weeklyData': weekly, 'dailyTrend': trend
        },
        'reports': reports
    }

    # Cache
    db_execute(db,
        'INSERT OR REPLACE INTO monthly_cache (year, month, data, updated_at) VALUES (?, ?, ?, CURRENT_TIMESTAMP)',
        (year, month, json.dumps(result, ensure_ascii=False))
    )
    db.commit()

    return jsonify(result)

@app.route('/api/reports/dates', methods=['GET'])
@login_required
def api_dates():
    db = get_db()
    rows = db_execute(db, 'SELECT DISTINCT date FROM daily_reports ORDER BY date DESC').fetchall()
    return jsonify({'dates': [r['date'] for r in rows]})

@app.route('/api/reports/stats', methods=['GET'])
@boss_required
def api_stats():
    db = get_db()
    total = db_execute(db, 'SELECT COUNT(*) as c FROM daily_reports').fetchone()['c']
    income = db_execute(db, 'SELECT SUM(total_income) as s FROM daily_reports').fetchone()['s'] or 0
    latest = db_execute(db, 'SELECT MAX(date) as d FROM daily_reports').fetchone()['d']
    return jsonify({'totalReports': total, 'totalIncome': income, 'latestDate': latest})

# --- 
# --- Excel Export -------------------------------------------------------------
# 酒店房型配置（固定库存）
ROOM_TYPE_INVENTORY = {
    "特惠大床房":   {"total": 2},
    "精选大床房":   {"total": 12},
    "豪华大床房":   {"total": 17},
    "城堡亲子房":  {"total": 4},
    "麻将双床房":   {"total": 7},
    "麻将套房":     {"total": 2},
    "圆床房":       {"total": 1},
}

@app.route("/api/reports/export/daily", methods=["GET"])
@boss_required
def api_export_daily():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 未安装"}), 500

    year = request.args.get("year")
    month = request.args.get("month")
    db = get_db()
    if year and month:
        rows = db_execute(db,
            "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date ASC",
            (f"{year}-{month.zfill(2)}%",)
        ).fetchall()
    else:
        rows = db_execute(db, "SELECT * FROM daily_reports ORDER BY date ASC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "日报明细"

    HDR_FILL = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUB_FILL = PatternFill("solid", fgColor="E8EDF5")
    SUB_FONT = Font(bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=14, color="1A1A2E")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    money_cols = {4, 6, 8, 10, 12, 14, 16, 18, 19, 24}

    ws.merge_cells("A1:X1")
    title_text = "华悦艺术酒店（冠华店）日报明细"
    if year and month:
        title_text += f" {year}年{int(month):02d}月"
    ws["A1"] = title_text
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws.row_dimensions[1].height = 28

    headers = [
        "日期", "班次",
        "美团间", "美团收入", "携程间", "携程收入",
        "飞猪间", "飞猪收入", "抖音间", "抖音收入",
        "微信间", "微信收入", "现金间", "现金收入",
        "支付宝间", "支付宝收入",
        "停车票", "停车收入", "税",
        "开房间", "入住率%", "RevPGR", "营业额",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = center; cell.border = border
    ws.row_dimensions[3].height = 22

    for ri, r in enumerate(rows, 4):
        vals = [
            r["date"], r["shift"] or "",
            r["meituan_rooms"], r["meituan_income"],
            r["ctrip_rooms"], r["ctrip_income"],
            r["fliggy_rooms"], r["fliggy_income"],
            r["douyin_rooms"], r["douyin_income"],
            r["wechat_rooms"], r["wechat_income"],
            r["cash_rooms"], r["cash_income"],
            r["alipay_rooms"], r["alipay_income"],
            r["parking_tickets"], r["parking_income"], r["tax"],
            r["total_rooms"], r["occupancy_rate"], r["revpgr"], r["total_income"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = border
            if col == 1:
                cell.value = r["date"]
                cell.font = Font(bold=True, color="0F3460")
                cell.alignment = center
            elif col == 2:
                cell.value = r["shift"] or ""
                cell.alignment = center
            elif col in money_cols:
                cell.value = round(float(v or 0), 2)
                cell.number_format = '"¥"#,##0.00'
                cell.font = Font(color="E94560", bold=True)
                cell.alignment = center
            elif col == 21:
                occ = float(v or 0)
                cell.value = round(occ, 1)
                cell.number_format = "0.0"
                cell.alignment = center
                cell.font = Font(color="2E7D32" if occ >= 70 else "E65100", bold=True)
            else:
                cell.value = v if v is not None else 0
                cell.alignment = center

    last_row = 4 + len(rows)
    ws.cell(row=last_row, column=1, value="汇总").fill = SUB_FILL
    ws.cell(row=last_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=last_row, column=1).alignment = center; ws.cell(row=last_row, column=1).border = border
    sum_cols = {3: "meituan_rooms", 4: "meituan_income", 5: "ctrip_rooms", 6: "ctrip_income",
                 7: "fliggy_rooms", 8: "fliggy_income", 9: "douyin_rooms", 10: "douyin_income",
                 11: "wechat_rooms", 12: "wechat_income", 13: "cash_rooms", 14: "cash_income",
                 15: "alipay_rooms", 16: "alipay_income", 17: "parking_tickets", 18: "parking_income",
                 19: "tax", 20: "total_rooms", 23: "total_income"}
    for col, key in sum_cols.items():
        val = sum(float(r[key] or 0) for r in rows)
        cell = ws.cell(row=last_row, column=col, value=round(val, 2))
        cell.fill = SUB_FILL; cell.font = SUB_FONT
        cell.alignment = center; cell.border = border
        if col in money_cols:
            cell.number_format = '"¥"#,##0.00'

    for i, w in enumerate([13, 7, 7, 12, 7, 12, 7, 12, 7, 12, 7, 12, 7, 12, 8, 12, 8, 12, 8, 8, 8, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"日报明细_{year or '全部'}_{month or ''}".strip("_") + ".xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/api/reports/export/roomtypes", methods=["GET"])
@boss_required
def api_export_roomtypes():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 未安装"}), 500

    year = request.args.get("year")
    month = request.args.get("month")
    db = get_db()
    if year and month:
        rows = db_execute(db,
            "SELECT * FROM daily_reports WHERE date LIKE ? ORDER BY date ASC",
            (f"{year}-{month.zfill(2)}%",)
        ).fetchall()
    else:
        rows = db_execute(db, "SELECT * FROM daily_reports ORDER BY date ASC").fetchall()

    reports = []
    for r in rows:
        d = dict(r)
        if d.get("room_types"):
            try:
                d["room_types"] = json.loads(d["room_types"])
            except:
                d["room_types"] = {}
        else:
            d["room_types"] = {}
        reports.append(d)

    wb = Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
    SUB_FILL  = PatternFill("solid", fgColor="E8EDF5")
    SUB_FONT  = Font(bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=14, color="1A1A2E")
    GOOD_FILL = PatternFill("solid", fgColor="E8F5E9")
    WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
    BAD_FILL  = PatternFill("solid", fgColor="FFEBEE")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    period = f"{year}年{int(month):02d}月" if (year and month) else "全部时间"
    # 过滤有房型数据的日报
    reports_with_room_data = [r for r in reports if r.get("room_types") and len(r.get("room_types", {})) > 0]
    valid_days = len(reports_with_room_data)
    total_days = len(reports)
    no_data_days = total_days - valid_days
    days = max(valid_days, 1)
    # 无房型数据的日期列表
    no_data_dates = [r.get("date", "") for r in reports if not r.get("room_types") or len(r.get("room_types", {})) == 0]
    no_data_dates.sort()
    no_data_dates_str = f"（日报1-{no_data_dates[0][-2:]}日无房型数据）" if no_data_dates else ""
    # 数据截至日期
    latest_date = max([r.get("date", "") for r in reports], default="")
    data_date_str = f"数据截至{latest_date}，" if latest_date else ""

    # Sheet 1: 房型汇总
    ws1 = wb.active
    ws1.title = "房型汇总"
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"华悦艺术酒店（冠华店）房型经营汇总  {period}"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = center
    ws1["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws1.row_dimensions[1].height = 28

    hdrs = ["房型", "总房间数", "累计开房数", "开房率", "日均开房(天数)", "总营业额", "平均房价", "开房率", "经营评价"]
    for col, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = center; cell.border = border
    ws1.row_dimensions[3].height = 22

    totals = {name: {"rooms": 0, "income": 0, "days": 0} for name in ROOM_TYPE_INVENTORY}
    for r in reports_with_room_data:
        rt = r.get("room_types") or {}
        for name in ROOM_TYPE_INVENTORY:
            if name in rt:
                rooms_val = int(rt[name].get("rooms", 0))
                totals[name]["rooms"]  += rooms_val
                totals[name]["income"] += float(rt[name].get("income", 0))
                if rooms_val > 0:
                    totals[name]["days"] += 1

    total_inv = sum(v["total"] for v in ROOM_TYPE_INVENTORY.values())
    total_occ = sum(v["rooms"] for v in totals.values())
    total_inc = sum(v["income"] for v in totals.values())

    for ri, (name, inv) in enumerate(ROOM_TYPE_INVENTORY.items(), 4):
        t = totals[name]
        occ = t["rooms"]; inc = t["income"]
        rt_days = t["days"]  # 该房型有开房的天数（保留用于其他参考）
        # 开房率：所有房型统一使用有房型数据的总天数作为分母
        occ_rate = (occ / (inv["total"] * valid_days) * 100) if valid_days > 0 else 0
        avg_p = inc / occ if occ > 0 else 0
        # 日均开房：所有房型统一使用有房型数据的总天数作为分母
        avg_per_day = occ / valid_days if valid_days > 0 else 0
        if occ == 0:
            rating, rfill = "无数据", BAD_FILL
        elif occ_rate >= 80:
            rating, rfill = "优秀", GOOD_FILL
        elif occ_rate >= 50:
            rating, rfill = "良好", PatternFill("solid", fgColor="E3F2FD")
        else:
            rating, rfill = "待提升", WARN_FILL

        vals = [name, inv["total"], occ, f"{occ_rate:.1f}%",
                f"{avg_per_day:.1f}天({valid_days}天)", f"¥{inc:,.2f}", f"¥{avg_p:.2f}", f"{occ_rate:.1f}%", rating]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=col, value=v)
            cell.border = border; cell.alignment = center
            if col == 9:
                cell.fill = rfill; cell.font = Font(bold=True)
            elif col == 1:
                cell.font = Font(bold=True, color="0F3460")
            elif col in (2, 3):
                cell.font = Font(bold=True)
        ws1.row_dimensions[ri].height = 20

    last = 4 + len(ROOM_TYPE_INVENTORY)
    ws1.cell(row=last, column=1, value="合计").fill = SUB_FILL
    ws1.cell(row=last, column=1).font = Font(bold=True, size=11)
    ws1.cell(row=last, column=1).alignment = center; ws1.cell(row=last, column=1).border = border
    ws1.cell(row=last, column=2, value=total_inv).fill = SUB_FILL
    ws1.cell(row=last, column=2).font = Font(bold=True); ws1.cell(row=last, column=2).alignment = center; ws1.cell(row=last, column=2).border = border
    ws1.cell(row=last, column=3, value=total_occ).fill = SUB_FILL
    ws1.cell(row=last, column=3).font = Font(bold=True); ws1.cell(row=last, column=3).alignment = center; ws1.cell(row=last, column=3).border = border
    ws1.cell(row=last, column=6, value=f"¥{total_inc:,.2f}").fill = SUB_FILL
    ws1.cell(row=last, column=6).font = Font(bold=True, color="E94560"); ws1.cell(row=last, column=6).alignment = center; ws1.cell(row=last, column=6).border = border
    avg_all = total_inc / total_occ if total_occ > 0 else 0
    ws1.cell(row=last, column=7, value=f"¥{avg_all:.2f}").fill = SUB_FILL
    ws1.cell(row=last, column=7).font = Font(bold=True); ws1.cell(row=last, column=7).alignment = center; ws1.cell(row=last, column=7).border = border
    for i, w in enumerate([14, 10, 12, 10, 16, 14, 12, 10, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A4"

    # Sheet 2: 每日明细（仅显示有房型数据的日报）
    ws2 = wb.create_sheet("每日房型明细")
    ws2.merge_cells("A1:K1")
    ws2["A1"] = f"华悦艺术酒店（冠华店）每日房型明细  {period}"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = center
    ws2["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws2.row_dimensions[1].height = 28

    daily_hdrs = ["日期", "班次"] + list(ROOM_TYPE_INVENTORY.keys()) + ["当日总开房", "当日营业额"]
    for col, h in enumerate(daily_hdrs, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = center; cell.border = border
    ws2.row_dimensions[3].height = 22

    # 仅显示有房型数据的日报
    for ri, r in enumerate(reports_with_room_data, 4):
        rt = r.get("room_types") or {}
        row_vals = [r["date"], r["shift"] or ""]
        day_total = 0
        for name in ROOM_TYPE_INVENTORY:
            rooms = int(rt.get(name, {}).get("rooms", 0))
            row_vals.append(rooms); day_total += rooms
        row_vals.append(day_total)
        row_vals.append(round(float(r.get("total_income", 0) or 0), 2))
        for col, v in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=col, value=v)
            cell.border = border; cell.alignment = center
            if col == 1:
                cell.font = Font(bold=True, color="0F3460")
            elif col == 11:
                cell.number_format = '"¥"#,##0.00'
                cell.font = Font(color="E94560", bold=True)
            elif 3 <= col <= 9:
                occ = int(v or 0)
                cell.fill = PatternFill("solid", fgColor="E8F5E9") if occ > 0 else PatternFill("solid", fgColor="FAFAFA")
                if occ == 0:
                    cell.font = Font(color="CCCCCC")

    for i, w in enumerate([13, 7, 10, 10, 10, 10, 10, 10, 10, 12, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "C4"

    # Sheet 3: 经营分析
    ws3 = wb.create_sheet("经营分析")
    ws3.merge_cells("A1:B1")
    ws3["A1"] = f"房型经营分析报告  {period}"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = center
    ws3["A1"].fill = PatternFill("solid", fgColor="F0F2F7")
    ws3.row_dimensions[1].height = 28
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 75

    # 重新计算分析数据（所有房型统一使用有房型数据的总天数作为分母）
    analyses = []
    for name, inv in ROOM_TYPE_INVENTORY.items():
        t = totals[name]
        occ = t["rooms"]; inc = t["income"]
        rt_days = t["days"]  # 保留参考
        # 开房率：所有房型统一使用有房型数据的总天数作为分母
        occ_rate = (occ / (inv["total"] * valid_days) * 100) if valid_days > 0 else 0
        avg_p = inc / occ if occ > 0 else 0
        analyses.append({"name": name, "inv": inv["total"], "occ": occ,
                          "inc": inc, "rate": occ_rate, "avg": avg_p, "days": valid_days})

    best_inc  = max([x for x in analyses if x["inc"] > 0], key=lambda x: x["inc"], default=None)
    best_rate = max([x for x in analyses if x["occ"] > 0], key=lambda x: x["rate"], default=None)
    worst = min([x for x in analyses if x["occ"] > 0], key=lambda x: x["rate"], default=None)
    zero_room_types = [x["name"] for x in analyses if x["occ"] == 0]
    low_rate_room_types = [x["name"] for x in analyses if x["occ"] > 0 and x["rate"] < 30]

    rows3 = [
        ("数据说明", f"{data_date_str}统计周期：{period}  |  月报共 {total_days} 天  |  有房型数据：{valid_days} 天  |  无房型数据：{no_data_days} 天{no_data_dates_str}（已排除）"),
        ("汇总", f"累计开房：{total_occ} 间夜  |  房型总收入：¥{total_inc:,.2f}  |  综合均价：{'¥'+str(round(total_inc/total_occ,2)) if total_occ > 0 else '¥0'}"),
        ("收入冠军", f'{best_inc["name"]}：总收入 ¥{best_inc["inc"]:,.2f}，累计 {best_inc["occ"]} 间夜，均价 ¥{best_inc["avg"]:.2f}。该房型表现优异，建议重点维护。'),
        ("开房率冠军", f'{best_rate["name"]}：开房率 {best_rate["rate"]:.1f}%，累计 {best_rate["occ"]} 间夜。{"接近满房，可考虑适当提价。" if best_rate["rate"] >= 80 else "表现良好，可作为定价参考。"}'),
        ("待提升房型", f'{worst["name"] if worst else "无"}：开房率 {worst["rate"]:.1f}%。建议：①检查该房型定价是否偏高；②优化线上展示图片和描述；③考虑是否改为长租或协议价。'),
    ]

    if zero_room_types:
        rows3.append(("空白房型", f'{",".join(zero_room_types)} 本月无订单。建议：①检查是否在OTA平台上架；②考虑合并至其他房型或调整配置；③如为主力房型，需重点排查原因。'))

    if low_rate_room_types:
        rows3.append(("低入住率房型", f'{",".join(low_rate_room_types)} 入住率低于30%。建议分析竞争对手同类房型定价，适当调低价格或增加套餐附加值（如含早、含门票等）。'))

    rows3.append(("经营建议", "①高收入房型保障流量支持，持续优化服务和口碑；②低入住率房型建议适度调价或优化套餐组合；③关注周末与工作日的需求差异，动态调整定价策略；④定期分析各渠道贡献占比，优化渠道资源配置。"))

    for ri, (label, content_txt) in enumerate(rows3, 3):
        c1 = ws3.cell(row=ri, column=1, value=label)
        c1.font = Font(bold=True, size=11, color="0F3460")
        c1.fill = PatternFill("solid", fgColor="E8EDF5")
        c1.alignment = center; c1.border = border
        c2 = ws3.cell(row=ri, column=2, value=content_txt)
        c2.font = Font(size=11)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        c2.border = border
        ws3.row_dimensions[ri].height = 32

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"房型汇总_{year or '全部'}_{month or ''}".strip("_") + ".xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


# --- Frontend ----------------------------------------------------------------
@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/<page>.html')
def page(page):
    return send_from_directory('public', f'{page}.html')

# --- Seed historical data ---------------------------------------------------
@app.route('/api/seed', methods=['POST'])
def api_seed():
    """预填历史数据（仅演示用，生产环境应删除此路由）"""
    db = get_db()
    seed_data = [
        {'date':'2026-04-01','shift':'早班','meituan_rooms':3,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':3,'meituan_income':576,'ctrip_income':416,'fliggy_income':208,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':624,'parking_tickets':4,'parking_income':60,'tax':138,'total_rooms':25,'avg_price':148,'occupancy_rate':64.0,'revpgr':94.7,'total_income':3485},
        {'date':'2026-04-02','shift':'早班','meituan_rooms':4,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':5,'cash_rooms':2,'alipay_rooms':3,'meituan_income':768,'ctrip_income':416,'fliggy_income':198,'douyin_income':396,'wechat_income':990,'cash_income':416,'alipay_income':624,'parking_tickets':3,'parking_income':45,'tax':132,'total_rooms':25,'avg_price':145,'occupancy_rate':76.0,'revpgr':110.2,'total_income':3808},
        {'date':'2026-04-03','shift':'早班','meituan_rooms':5,'ctrip_rooms':3,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':3,'alipay_rooms':4,'meituan_income':960,'ctrip_income':624,'fliggy_income':198,'douyin_income':396,'wechat_income':792,'cash_income':624,'alipay_income':828,'parking_tickets':5,'parking_income':75,'tax':156,'total_rooms':25,'avg_price':148,'occupancy_rate':88.0,'revpgr':130.2,'total_income':4422},
        {'date':'2026-04-04','shift':'早班','meituan_rooms':5,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':5,'cash_rooms':3,'alipay_rooms':3,'meituan_income':960,'ctrip_income':624,'fliggy_income':396,'douyin_income':198,'wechat_income':990,'cash_income':624,'alipay_income':624,'parking_tickets':6,'parking_income':90,'tax':152,'total_rooms':25,'avg_price':145,'occupancy_rate':88.0,'revpgr':127.6,'total_income':4416},
        {'date':'2026-04-05','shift':'早班','meituan_rooms':4,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':5,'cash_rooms':3,'alipay_rooms':2,'meituan_income':768,'ctrip_income':624,'fliggy_income':396,'douyin_income':198,'wechat_income':990,'cash_income':624,'alipay_income':416,'parking_tickets':5,'parking_income':75,'tax':144,'total_rooms':25,'avg_price':145,'occupancy_rate':80.0,'revpgr':116.0,'total_income':4016},
        {'date':'2026-04-06','shift':'早班','meituan_rooms':5,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':3,'meituan_income':960,'ctrip_income':624,'fliggy_income':396,'douyin_income':396,'wechat_income':792,'cash_income':416,'alipay_income':624,'parking_tickets':5,'parking_income':75,'tax':148,'total_rooms':25,'avg_price':146,'occupancy_rate':84.0,'revpgr':122.6,'total_income':4208},
        {'date':'2026-04-07','shift':'早班','meituan_rooms':5,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':5,'cash_rooms':2,'alipay_rooms':4,'meituan_income':960,'ctrip_income':416,'fliggy_income':198,'douyin_income':396,'wechat_income':990,'cash_income':416,'alipay_income':828,'parking_tickets':4,'parking_income':60,'tax':136,'total_rooms':25,'avg_price':148,'occupancy_rate':84.0,'revpgr':124.3,'total_income':4204},
        {'date':'2026-04-08','shift':'早班','meituan_rooms':4,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':2,'meituan_income':768,'ctrip_income':624,'fliggy_income':396,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':416,'parking_tickets':3,'parking_income':45,'tax':124,'total_rooms':25,'avg_price':145,'occupancy_rate':72.0,'revpgr':104.4,'total_income':3610},
        {'date':'2026-04-09','shift':'早班','meituan_rooms':3,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':1,'wechat_rooms':3,'cash_rooms':2,'alipay_rooms':2,'meituan_income':576,'ctrip_income':416,'fliggy_income':198,'douyin_income':198,'wechat_income':594,'cash_income':416,'alipay_income':416,'parking_tickets':3,'parking_income':45,'tax':116,'total_rooms':25,'avg_price':142,'occupancy_rate':56.0,'revpgr':79.5,'total_income':2814},
        {'date':'2026-04-10','shift':'早班','meituan_rooms':3,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':2,'meituan_income':576,'ctrip_income':416,'fliggy_income':198,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':416,'parking_tickets':2,'parking_income':30,'tax':116,'total_rooms':25,'avg_price':145,'occupancy_rate':60.0,'revpgr':87.0,'total_income':3012},
        {'date':'2026-04-11','shift':'早班','meituan_rooms':4,'ctrip_rooms':2,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':2,'meituan_income':768,'ctrip_income':416,'fliggy_income':396,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':416,'parking_tickets':3,'parking_income':45,'tax':124,'total_rooms':25,'avg_price':145,'occupancy_rate':68.0,'revpgr':98.6,'total_income':3402},
        {'date':'2026-04-12','shift':'早班','meituan_rooms':3,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':3,'cash_rooms':2,'alipay_rooms':2,'meituan_income':576,'ctrip_income':416,'fliggy_income':198,'douyin_income':396,'wechat_income':594,'cash_income':416,'alipay_income':416,'parking_tickets':3,'parking_income':45,'tax':118,'total_rooms':25,'avg_price':142,'occupancy_rate':60.0,'revpgr':85.2,'total_income':3012},
        {'date':'2026-04-13','shift':'早班','meituan_rooms':4,'ctrip_rooms':3,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':3,'meituan_income':768,'ctrip_income':624,'fliggy_income':198,'douyin_income':396,'wechat_income':792,'cash_income':416,'alipay_income':624,'parking_tickets':4,'parking_income':60,'tax':138,'total_rooms':25,'avg_price':145,'occupancy_rate':76.0,'revpgr':110.2,'total_income':3818},
        {'date':'2026-04-14','shift':'早班','meituan_rooms':5,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':5,'cash_rooms':3,'alipay_rooms':3,'meituan_income':960,'ctrip_income':624,'fliggy_income':396,'douyin_income':198,'wechat_income':990,'cash_income':624,'alipay_income':624,'parking_tickets':6,'parking_income':90,'tax':152,'total_rooms':25,'avg_price':145,'occupancy_rate':88.0,'revpgr':127.6,'total_income':4416},
        {'date':'2026-04-15','shift':'早班','meituan_rooms':5,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':4,'meituan_income':960,'ctrip_income':416,'fliggy_income':198,'douyin_income':396,'wechat_income':792,'cash_income':416,'alipay_income':828,'parking_tickets':5,'parking_income':75,'tax':144,'total_rooms':25,'avg_price':148,'occupancy_rate':80.0,'revpgr':118.4,'total_income':4006},
        {'date':'2026-04-16','shift':'早班','meituan_rooms':4,'ctrip_rooms':2,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':3,'meituan_income':768,'ctrip_income':416,'fliggy_income':396,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':624,'parking_tickets':3,'parking_income':45,'tax':130,'total_rooms':25,'avg_price':145,'occupancy_rate':72.0,'revpgr':104.4,'total_income':3610},
        {'date':'2026-04-17','shift':'早班','meituan_rooms':4,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':2,'meituan_income':768,'ctrip_income':416,'fliggy_income':198,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':416,'parking_tickets':3,'parking_income':45,'tax':122,'total_rooms':25,'avg_price':145,'occupancy_rate':64.0,'revpgr':92.8,'total_income':3204},
        {'date':'2026-04-18','shift':'早班','meituan_rooms':3,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':1,'wechat_rooms':3,'cash_rooms':2,'alipay_rooms':2,'meituan_income':576,'ctrip_income':416,'fliggy_income':198,'douyin_income':198,'wechat_income':594,'cash_income':416,'alipay_income':416,'parking_tickets':2,'parking_income':30,'tax':112,'total_rooms':25,'avg_price':142,'occupancy_rate':56.0,'revpgr':79.5,'total_income':2814},
        {'date':'2026-04-19','shift':'早班','meituan_rooms':3,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':1,'meituan_income':576,'ctrip_income':416,'fliggy_income':198,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':208,'parking_tickets':2,'parking_income':30,'tax':108,'total_rooms':25,'avg_price':143,'occupancy_rate':56.0,'revpgr':80.1,'total_income':2804},
        {'date':'2026-04-20','shift':'早班','meituan_rooms':4,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':3,'meituan_income':768,'ctrip_income':416,'fliggy_income':198,'douyin_income':396,'wechat_income':792,'cash_income':416,'alipay_income':624,'parking_tickets':3,'parking_income':45,'tax':128,'total_rooms':25,'avg_price':145,'occupancy_rate':72.0,'revpgr':104.4,'total_income':3610},
        {'date':'2026-04-21','shift':'早班','meituan_rooms':4,'ctrip_rooms':3,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':5,'cash_rooms':2,'alipay_rooms':3,'meituan_income':768,'ctrip_income':624,'fliggy_income':198,'douyin_income':396,'wechat_income':990,'cash_income':416,'alipay_income':624,'parking_tickets':5,'parking_income':75,'tax':140,'total_rooms':25,'avg_price':145,'occupancy_rate':80.0,'revpgr':116.0,'total_income':4016},
        {'date':'2026-04-22','shift':'早班','meituan_rooms':5,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':2,'wechat_rooms':5,'cash_rooms':3,'alipay_rooms':3,'meituan_income':960,'ctrip_income':624,'fliggy_income':396,'douyin_income':396,'wechat_income':990,'cash_income':624,'alipay_income':624,'parking_tickets':5,'parking_income':75,'tax':150,'total_rooms':25,'avg_price':145,'occupancy_rate':88.0,'revpgr':127.6,'total_income':4414},
        {'date':'2026-04-23','shift':'早班','meituan_rooms':5,'ctrip_rooms':2,'fliggy_rooms':1,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':4,'meituan_income':960,'ctrip_income':416,'fliggy_income':198,'douyin_income':396,'wechat_income':792,'cash_income':416,'alipay_income':828,'parking_tickets':5,'parking_income':75,'tax':144,'total_rooms':25,'avg_price':148,'occupancy_rate':80.0,'revpgr':118.4,'total_income':4006},
        {'date':'2026-04-24','shift':'早班','meituan_rooms':4,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':1,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':3,'meituan_income':768,'ctrip_income':624,'fliggy_income':396,'douyin_income':198,'wechat_income':792,'cash_income':416,'alipay_income':624,'parking_tickets':4,'parking_income':60,'tax':136,'total_rooms':25,'avg_price':145,'occupancy_rate':76.0,'revpgr':110.2,'total_income':3818},
        {'date':'2026-04-25','shift':'早班','meituan_rooms':5,'ctrip_rooms':3,'fliggy_rooms':2,'douyin_rooms':2,'wechat_rooms':4,'cash_rooms':2,'alipay_rooms':4,'meituan_income':960,'ctrip_income':624,'fliggy_income':396,'douyin_income':396,'wechat_income':792,'cash_income':416,'alipay_income':828,'parking_tickets':5,'parking_income':75,'tax':150,'total_rooms':25,'avg_price':146,'occupancy_rate':88.0,'revpgr':128.5,'total_income':4412},
    ]

    for r in seed_data:
        db_execute(db, '''
          INSERT OR REPLACE INTO daily_reports
          (date, shift, meituan_rooms, ctrip_rooms, fliggy_rooms, douyin_rooms,
           wechat_rooms, cash_rooms, alipay_rooms, meituan_income, ctrip_income,
           fliggy_income, douyin_income, wechat_income, cash_income, alipay_income,
           parking_tickets, parking_income, tax, total_rooms, avg_price,
           occupancy_rate, revpgr, total_income, note, uploaded_by)
          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'seed')
        ''', (
          r['date'], r.get('shift',''), r.get('meituan_rooms',0), r.get('ctrip_rooms',0),
          r.get('fliggy_rooms',0), r.get('douyin_rooms',0), r.get('wechat_rooms',0),
          r.get('cash_rooms',0), r.get('alipay_rooms',0), r.get('meituan_income',0),
          r.get('ctrip_income',0), r.get('fliggy_income',0), r.get('douyin_income',0),
          r.get('wechat_income',0), r.get('cash_income',0), r.get('alipay_income',0),
          r.get('parking_tickets',0), r.get('parking_income',0), r.get('tax',0),
          r.get('total_rooms',0), r.get('avg_price',0), r.get('occupancy_rate',0),
          r.get('revpgr',0), r.get('total_income',0)
        ))
    db.commit()
    db_execute(db, 'DELETE FROM monthly_cache')
    db.commit()
    return jsonify({'success': True, 'count': len(seed_data)})

# --- Run ---------------------------------------------------------------------
if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 3000))
    print(f'[HOTEL] Server started: http://localhost:{port}')
    print('[STAFF] Login: 19128957480 (any 6-digit code)')
    print('[BOSS]  Login: 13802531098 / 18602032126')
    print('📋 预填数据: POST http://localhost:{port}/api/seed')
    app.run(host='0.0.0.0', port=port, debug=False)
